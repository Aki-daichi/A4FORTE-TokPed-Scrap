from openpyxl import Workbook, load_workbook

file = 'ExcelFile.xlsx'
wb = load_workbook(file)
# wb.create_sheet("BJIR") #Membuat Sheet baru
ws = wb['BJIR']  # Accessing sheet
print(ws['A1'].value)
ws['A1'] = "Asyam"
wb.save(file)

wb = Workbook()
ws = wb.active
ws.title = "DATASET" #mengganti nama current sheet
ws.append(['Nama Toko',"Waktu Pesanan di Proses", 'Rata-Rata Penilaian', "Jumlah Pemberi Nilai","Jumlah Penjualan"])
wb.save("Test2.xlsx")

from openpyxl import Workbook, load_workbook
import json
import re
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def get_digit_in_string(text):
    pattern = r'\d+'
    match = re.search(pattern, text)
    if match:
        # Extract the matched part and convert it to an integer
        number = int(match.group())
        return number
    else:
        return 0


def contains_sambal_or_sambel(text):
    # Regular expression pattern to match "sambal" or "sambel"
    pattern = r'\bsambal\b|\bsambel\b'
    # Search the text for the pattern
    match = re.search(pattern, text, re.IGNORECASE)
    # Regular expression pattern to match "sambal" or "sambel"
    pattern2 = r'\btempat\b|\bcup\b|\bwadah\b'
    # Search the text for the pattern
    match2 = re.search(pattern2, text, re.IGNORECASE)
    # Return True if the pattern is found, otherwise False
    return bool(match and (not match2))


def WaitById(driver, nameId):
    try:
        WebDriverWait(driver, 0.2).until(
            EC.presence_of_element_located((By.ID, nameId))
        )
        return driver.find_element(By.ID, nameId)
    except:
        driver.quit()
        exit()


def WaitByClass(driver, nameClass):
    try:
        wait = driver.find_elements(By.CLASS_NAME, nameClass)
        return wait
    except:
        print("Continue")


def WaitByClassSingle(driver, nameClass, timeout):
    try:
        wait = WebDriverWait(driver, timeout)
        wait = wait.until(EC.presence_of_element_located((By.CLASS_NAME, nameClass)))
        return wait
    except:
        return 0


def WaitByCSS(driver, nameCSS):
    try:
        wait = driver.find_element(By.CSS_SELECTOR, nameCSS)
        return wait
    except:
        driver.quit()
        exit()

"""
wb = Workbook()
ws = wb.active
ws.title = "DATASET"  # mengganti nama current sheet
ws.append(['Nama Toko', 'Nama Produk Terlaku #1', 'Harga Produk Terlaku #1', "Jumlah Penjualan Produk Terlaku #1",
           "Jumlah Pemberi Nilai Produk Terlaku #1",
           'Rata-Rata Penilaian Produk Terlaku #1', 'Nama Produk Terlaku #2', 'Harga Produk Terlaku #2',
           "Jumlah Penjualan Produk Terlaku #2",
           "Jumlah Pemberi Nilai Produk Terlaku #2", 'Rata-Rata Penilaian Produk Terlaku #2', 'Nama Produk Terlaku #3',
           'Harga Produk Terlaku #3',
           "Jumlah Penjualan Produk Terlaku #3", "Jumlah Pemberi Nilai Produk Terlaku #3",
           'Rata-Rata Penilaian Produk Terlaku #3', 'Nama Produk Terlaku #4', 'Harga Produk Terlaku #4',
           "Jumlah Penjualan Produk Terlaku #4",
           "Jumlah Pemberi Nilai Produk Terlaku #4", 'Rata-Rata Penilaian Produk Terlaku #4', 'Nama Produk Terlaku #5',
           'Harga Produk Terlaku #5',
           "Jumlah Penjualan Produk Terlaku #5", "Jumlah Pemberi Nilai Produk Terlaku #5",
           'Rata-Rata Penilaian Produk Terlaku #5'])
wb.save("HasilScrapingZaki.xlsx")
"""
driver = webdriver.Chrome()
driver.maximize_window()
i = 0
amount_of_toko = 293
list_link_toko = []
with open('link.txt', 'r') as file:
    for line in file:
        list_link_toko.append(line.strip())
for link_toko in list_link_toko:
    amount_of_toko = amount_of_toko + 1
    firstLink = link_toko
    curPage = 1
    WebLink = firstLink + "/product/page/" + str(curPage) + "?sort=8"
    driver.get(WebLink)
    HaveGeneral = True
    nama_toko = WaitByClassSingle(driver, "css-fzzhh3", 2).text
    print(nama_toko)
    driver.implicitly_wait(30)
    asa = 0
    aced = 100
    file = open(WebLink[26:(len(firstLink))], "a")
    file.close()
    list_link_produk = []
    if driver.current_url in firstLink:
        WebLink = firstLink + "/page/" + str(curPage) + "?sort=8"
        HaveGeneral = False
        print(WebLink)
        driver.get(WebLink)
    while True:
        time.sleep(4)
        while asa < 5000:
            driver.execute_script("window.scrollTo(" + str(asa) + ", " + str(aced) + ");")
            asa = asa + 100
            aced = aced + 100
            time.sleep(0.05)
        driver.implicitly_wait(2)
        try:
            EndOfPage = WaitByClassSingle(driver, "css-1x8eu1x-unf-heading", 1).text
            print(EndOfPage)
            break
        except:
            print("Continue")
        driver.implicitly_wait(30)
        produks = WaitByClass(driver, "css-19oqosi")
        for produk in produks:
            linkProduk = (produk.find_element(By.TAG_NAME, "a")).get_attribute("href")
            print(linkProduk.lower()[(len(firstLink) + 1):(len(linkProduk))])
            if contains_sambal_or_sambel(linkProduk.lower()[(len(firstLink) + 1):(len(linkProduk))]) is True:
                list_link_produk.append(linkProduk)
                i = i + 1
        print(i)
        time.sleep(1)
        curPage = curPage + 1
        asa = 0
        aced = 100
        if HaveGeneral is True:
            WebLink = firstLink + "/product/page/" + str(curPage) + "?sort=8"
        else:
            WebLink = firstLink + "/page/" + str(curPage) + "?sort=8"
        driver.get(WebLink)
        if i > 5 or curPage > 30:
            break
    i = 0
    ws_akan_di_append = [nama_toko]
    for produk in list_link_produk:
        i = i + 1
        if i > 5:
            i = 0
            break
        driver.implicitly_wait(50)
        driver.get(produk)
        driver.implicitly_wait(5)
        time.sleep(3)
        name_produks = WaitByClass(driver, "css-jmbq56")
        print(len(name_produks))
        name_produk = ""
        for nama in name_produks:
            try:
                print(nama.find_element(By.CLASS_NAME, "css-ga6qsf").text)
                name_produk = nama.find_element(By.CLASS_NAME, "css-ga6qsf").text
            except:
                i = i
        print(name_produk)
        ws_akan_di_append.append(name_produk)
        price_produk = get_digit_in_string(WaitByClassSingle(driver, "price", 2).text.replace('.', ''))
        print(str(price_produk))
        ws_akan_di_append.append(str(price_produk))
        bar_info = WaitByClass(driver, "css-vni7t6-unf-heading")
        total_terjual = ""
        total_pemberi_rating = ""
        mean_rating = ""
        if bar_info != 0:
            for info in bar_info:
                if "Terjual" in info.text:
                    total_terjual = get_digit_in_string(info.text)
                    if "rb" in info.text:
                        total_terjual = total_terjual * 1000
                    print(str(total_terjual))
                    total_terjual = str(total_terjual)
                elif "rating" in info.text:
                    total_pemberi_rating = get_digit_in_string(
                        info.text[12:len(info.text)].replace('.', '').replace('\n', ''))
                    print(str(total_pemberi_rating))
                    print(info.find_element(By.CLASS_NAME, "main").text)
                    total_pemberi_rating = str(total_pemberi_rating)
                    mean_rating = info.find_element(By.CLASS_NAME, "main").text
        ws_akan_di_append.append(total_terjual)
        ws_akan_di_append.append(total_pemberi_rating)
        ws_akan_di_append.append(mean_rating)
    if i <= 5:
        for j in range(i, 6):
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
            ws_akan_di_append.append("")
    if True:
        wbName = "HasilScrapingZaki.xlsx"
        wb = load_workbook(wbName)
        ws = wb.active
        ws['A'+str(amount_of_toko)] = ws_akan_di_append[0]
        ws['B'+str(amount_of_toko)] = ws_akan_di_append[1]
        ws['C'+str(amount_of_toko)] = ws_akan_di_append[2]
        ws['D'+str(amount_of_toko)] = ws_akan_di_append[3]
        ws['E'+str(amount_of_toko)] = ws_akan_di_append[4]
        ws['F'+str(amount_of_toko)] = ws_akan_di_append[5]
        ws['G'+str(amount_of_toko)] = ws_akan_di_append[6]
        ws['H'+str(amount_of_toko)] = ws_akan_di_append[7]
        ws['I'+str(amount_of_toko)] = ws_akan_di_append[8]
        ws['J'+str(amount_of_toko)] = ws_akan_di_append[9]
        ws['K'+str(amount_of_toko)] = ws_akan_di_append[10]
        ws['L'+str(amount_of_toko)] = ws_akan_di_append[11]
        ws['M'+str(amount_of_toko)] = ws_akan_di_append[12]
        ws['N'+str(amount_of_toko)] = ws_akan_di_append[13]
        ws['O'+str(amount_of_toko)] = ws_akan_di_append[14]
        ws['P'+str(amount_of_toko)] = ws_akan_di_append[15]
        ws['Q'+str(amount_of_toko)] = ws_akan_di_append[16]
        ws['R'+str(amount_of_toko)] = ws_akan_di_append[17]
        ws['S'+str(amount_of_toko)] = ws_akan_di_append[18]
        ws['T'+str(amount_of_toko)] = ws_akan_di_append[19]
        ws['U'+str(amount_of_toko)] = ws_akan_di_append[20]
        ws['V'+str(amount_of_toko)] = ws_akan_di_append[21]
        ws['W' + str(amount_of_toko)] = ws_akan_di_append[22]
        ws['X' + str(amount_of_toko)] = ws_akan_di_append[23]
        ws['Y' + str(amount_of_toko)] = ws_akan_di_append[24]
        ws['Z' + str(amount_of_toko)] = ws_akan_di_append[25]
        wb.save("HasilScrapingZaki.xlsx")
        print("TOKO KE "+str(amount_of_toko))
        i = 0
    list_link_produk.clear()
    ws.append(ws_akan_di_append)
print()

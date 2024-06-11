from openpyxl import Workbook, load_workbook

list_link_toko = []
with open('link.txt', 'r') as file:
    for line in file:
        list_link_toko.append(line.strip())
print(list_link_toko)

wbName = "TokoSambal.xlsx"
wb = load_workbook(wbName)
ws = wb.active
amountOfShopAdded = 0
list_unik_toko = []
for i in range(2, 1083):
    if ws['A' + str(i)].value not in list_link_toko:
        list_unik_toko.append(ws['A' + str(i)].value)
        print(str(i)+'\n')
        amountOfShopAdded += 1
        if amountOfShopAdded == 530:
            break
with open('link.txt', 'a') as file:
    for each in list_unik_toko:
        file.write(each+"\n")
